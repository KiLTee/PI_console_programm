# import json
# import openpyxl
import random
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl import Workbook
from os import path
from datetime import datetime
from test import find_patch


# Парсинг базы БДУ (по критериям)
def parse_BDU():
    wb = load_workbook("./Excel/vullist.xlsx")
    BDU = {}
    mass = []
    sheet = wb.get_sheet_by_name("Sheet")
    vendor_to_find = "VMware Inc."
    date_to_find = "2023"
    i = 2
    while isinstance(sheet.cell(row=i, column=1).value, str):
        if isinstance(sheet.cell(row=i, column=4).value, str) and isinstance(
            sheet.cell(row=i, column=10).value, str
        ):
            vendor_names = sheet.cell(row=i, column=4).value.split(", ")
            dates = sheet.cell(row=i, column=10).value.split(".")

            for j in range(len(vendor_names)):
                if vendor_names[j] == vendor_to_find and dates[2] == date_to_find:

                    for q in range(1, 23):
                        mass.append(sheet.cell(row=i, column=q).value)
                    res = mass
                    BDU[sheet.cell(row=i, column=1).value] = res
                    mass = []
        i += 1
    print(BDU)
    #return BDU


def xlsx(KLA_path, CVE_list_path, BDU_FSTEK_path):
    def open_sheet(book_path, sheet_name):  # - Функция открытия страницы в книге excel
        wb = load_workbook(book_path)  # - Открываем книгу
        return wb[sheet_name]  # - Возварщаем открытую страницу

    sheet = open_sheet(KLA_path, "Details")  # - Открытие страницы excel
    i = 2
    result = {}  # - Словарь для хранения результатов обработки Отчета об уязвимостях
    while isinstance(
        sheet.cell(row=i, column=1).value, str
    ):  # - До тех пор, пока тип данных в ячейке - строка
        result[sheet.cell(row=i, column=2).value] = [
            sheet.cell(row=i, column=j).value
            for j in range(1, 18)
            if sheet.cell(row=i, column=1).value == "Критическое"
        ]
        i += 1  # - для счетчика устроуйств не хватает 0 в конце списка

    # - Обработка списка уязвимотсей (CVE List) от Касперского (берем рекомендованные пачти)
    sheet = open_sheet(CVE_list_path, "CveList")
    patch_names = {}
    i = 2
    while isinstance(
        sheet.cell(row=i, column=1).value, str
    ):  # - Делаем выборку рекомендованных патчей от Касперского
        for key in result:
            if sheet.cell(row=i, column=1).value == key:
                patch_names[key] = [
                    sheet.cell(row=i, column=4).value,
                    sheet.cell(row=i, column=5).value,
                ]
        i += 1

    # - Обработка базы ФСТЭК
    sheet = open_sheet(BDU_FSTEK_path, "Sheet")
    BDU = {}  # - Словарь для хранения результатов обработки базы БДУ ФСТЭК
    mass = []  # - Массив для переноса значений в словарь
    i = 2
    while isinstance(sheet.cell(row=i, column=1).value, str):
        for key in result:
            cve_name = sheet.cell(row=i, column=19).value
            if isinstance(cve_name, str) and cve_name.split(", ")[0] == result[key][16]:
                for j in range(1, 23):
                    mass.append(sheet.cell(row=i, column=j).value)
                res = mass
                BDU[key] = res
            mass = []
        i += 1

    for key in BDU:  # - Формирование вывода графы "Меры устарнения"
        m = find_patch(
            BDU[key][3], BDU[key][4], [patch_names[key][0], patch_names[key][1]]
        )  # - вызов обработки КС РПС
        print(key, m)
        BDU[key][13] = (
            "Возможные меры по устранению, согласно БДУ ФСТЭК:\n"
            + str(BDU[key][13])
            + "\nВозможные меры по устранению, согласно KLA:\n"
            + "Рекомендованный основной патч:\n"
            + str(patch_names[key][0])
            if str(patch_names[key][0]) != None
            else (
                "-"
                + "\nРекомендованый дополнительный патч для закрытия: "
                + "\n"
                + str(patch_names[key][1])
                if str(patch_names[key][1]) != None
                else "-"
            )
        )
        BDU[key][3] = m
        print("\n" + BDU[key][3])
    print(BDU)
    return BDU


def make_base(BDU, author_name):
    def feel_cells(sheet, key, i, author_name):
        sheet.cell(row=i, column=1).value = key
        sheet.cell(row=i, column=2).value = (
            str(datetime.now().day)
            + "."
            + str(datetime.now().month)
            + "."
            + str(datetime.now().year)
        )
        sheet.cell(row=i, column=3).value = (
            str(datetime.now().hour) + ":" + str(datetime.now().minute)
        )
        sheet.cell(row=i, column=4).value = author_name
        sheet.cell(row=i, column=5).hyperlink = (
            f"./Результаты/Паспорт уязвимости {key}.docx"
        )
        sheet.cell(row=i, column=6).value = BDU[key][2]
        sheet.cell(row=i, column=7).value = BDU[key][12]
        sheet.cell(row=i, column=8).value = random.randint(
            1, 16
        )  # ЗАМЕНИТЬ ПОТОМ НА НОРАЛЬНЫЙ СЧЕТЧИК

    if path.exists("./База.xlsx"):
        wb = load_workbook("./База.xlsx")
        sheet = wb["База"]
        i = 2
        while isinstance(sheet.cell(row=i, column=1).value, str):
            for key in BDU:
                if key == sheet.cell(row=i, column=1).value:
                    BDU.pop(key)
            i += 1
        print(BDU)

        if len(BDU) > 0:
            for key in BDU:
                feel_cells(sheet, key, i, author_name)
                i += 1

        wb.save("./База.xlsx")
    else:
        wb = Workbook().active
        sheet = wb["Sheet"]
        sheet.title = "База"
        Workbook().save("./База.xlsx")
        wb = load_workbook("./База.xlsx")

        sheet.cell(row=1, column=1).value = "Уязвимость"
        sheet.cell(row=1, column=2).value = "Дата создания отчета"
        sheet.cell(row=1, column=3).value = "Время создания отчета"
        sheet.cell(row=1, column=4).value = "Кем создан отчет"
        sheet.cell(row=1, column=5).value = "Путь к файлу"
        sheet.cell(row=1, column=6).value = "Описание"
        sheet.cell(row=1, column=7).value = "Критичность"
        sheet.cell(row=1, column=8).value = "Количество устройств"
        wb.save("./База.xlsx")

        i = 2
        for key in BDU:
            feel_cells(sheet, key, i, author_name)
            i += 1

        Workbook().save("./База.xlsx")


# -----------Прототипы-----------
"""
def parse_json(file_path):
    with open(file_path, 'r') as f:
        data = json.load(f)
        return data

def print_cve_info_json(cve_info):
    print('ID: ', cve_info['cveMetadata']['cveId'])
    print('Продукт: ', cve_info["containers"]["cna"]["affected"][0]["product"])
    print('Вендор: ', cve_info['containers']['cna']['affected'][0]['vendor'])
    print('Дата публикации: ', cve_info['cveMetadata']['datePublished'])
    print('Платформы: ', cve_info['containers']['cna']['affected'][0]['platforms'])
    print('Версия CVSS: ', cve_info['containers']['cna']['metrics'][0]['cvssV3_1']['baseScore'])
    print('И тд') 

def parse_xml(file_path):
    with open(file_path, 'r') as f:
        tree = ET.parse(f)
        root = tree.getroot()

    # Получаем список всех дочерних элементов корневого элемента
    children = root.iter('*')

    # Выводим содержимое каждого дочернего элемента на экран
    for child in children:
        print(child.tag, child.text)

    # Закрываем файл
    f.close()

def print_cve_info_xml():
    pass
"""
